import openpyxl, json, re, os

def clean_answers(sol):
    """Produce list of normalised accepted answer strings."""
    if sol is None:
        return []
    sol = str(sol).strip()
    parts = re.split(r'\s+or\s+', sol, flags=re.IGNORECASE)
    accepted = []
    for p in parts:
        raw = p.strip()
        # strip LaTeX $...$
        raw = re.sub(r'\$', '', raw)
        # strip LaTeX commands like \times \circ \approx \log etc.
        raw = re.sub(r'\\[a-zA-Z]+\s*', '', raw)
        # strip curly braces
        raw = re.sub(r'[{}]', '', raw)
        # strip trailing units (players type bare number)
        stripped = re.sub(
            r'\s*(mph|km\/h|m\/s|days?|hours?|hrs?|mins?|minutes?|seconds?|secs?|%|°)\s*$',
            '', raw, flags=re.IGNORECASE).strip()
        for candidate in {raw.strip().lower(), stripped.lower()}:
            if candidate and candidate not in accepted:
                accepted.append(candidate)
    return accepted

def first_sentence(text):
    """Return first sentence from reasoning for a short hint."""
    if not text:
        return ''
    m = re.match(r'^(.+?[.!?])\s', text + ' ')
    return m.group(1).strip() if m else text.strip()

def process_sheet(ws):
    problems = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ptype, problem, reasoning, solution = (row + (None,)*4)[:4]
        if not problem:
            continue
        problems.append({
            'type':     str(ptype or '').strip(),
            'problem':  str(problem).strip(),
            'hint':     str(reasoning or '').strip(),
            'solution': str(solution or '').strip(),
            'answers':  clean_answers(solution),
        })
    return problems

LORE = [
    {
        'title':   'The Birth of Zero',
        'content': 'In 7th-century India, mathematician Brahmagupta defined zero as the result of subtracting a number from itself. The concept of śūnya — nothingness as a quantity — was revolutionary. Before this, no civilisation could mathematically represent the absence of something. This single idea made algebra, calculus, and all of modern computing possible.',
        'author':  '— Brahmagupta, Brahmasphutasiddhānta (628 AD)'
    },
    {
        'title':   "Euclid's Infinite Primes",
        'content': 'Written around 300 BC, Euclid proved that infinitely many primes exist with a single elegant argument. Assume a finite list. Multiply them all and add 1. The result is either prime, or has a prime factor not in your list — contradiction either way. It remains the first proof taught to every mathematics student on Earth.',
        'author':  '— Euclid, Elements Book IX (c. 300 BC)'
    },
   {
        'title':   'Archimedes and the Value of π',
        'content': 'Without a calculator, Archimedes calculated π to stunning accuracy by drawing 96-sided polygons inside and outside a circle. He proved π lies between 3 + 10⁄71 and 3 + 1⁄7. His method — exhaustion by polygons — remained the best technique for computing π for nearly two thousand years.',
        'author':  '— Archimedes of Syracuse (287–212 BC)'
    },
    {
        'title':   'The Child Who Summed to 5050',
        'content': 'At age 10, Carl Friedrich Gauss was told to sum every integer from 1 to 100 as a classroom punishment. Within seconds he wrote 5050. He realised that pairs from each end (1+100, 2+99 …) each equal 101, and there are exactly 50 such pairs. The formula n(n+1)⁄2 — which you just used — was discovered by a bored schoolboy who became one of history\'s greatest mathematicians.',
        'author':  '— Carl Friedrich Gauss (1777–1855)'
    },
]

LEVEL_META = [
    {'id':1,'name':'The Foundation',        'subtitle':'Begin Your Journey',    'timer':None,'badge':'Apprentice Scholar','badgeIcon':'🥉','minScore':60},
    {'id':2,'name':'The Tactician',          'subtitle':'Sharpen Your Mind',     'timer':60,  'badge':'Geometric Mind',     'badgeIcon':'🥈','minScore':65},
    {'id':3,'name':'The Strategist',         'subtitle':'Command the Numbers',   'timer':45,  'badge':'Algebraic Sage',     'badgeIcon':'🥇','minScore':70},
    {'id':4,'name':'The Archimedean Master', 'subtitle':'Claim Your Legacy',     'timer':30,  'badge':'Archimedean Master', 'badgeIcon':'💎','minScore':75},
]

os.makedirs('js', exist_ok=True)

wb = openpyxl.load_workbook('Logic_and_Arithmetic_Game_Problems.xlsx')
levels = []
for i, sheet in enumerate(wb.sheetnames):
    meta = LEVEL_META[i].copy()
    meta['problems'] = process_sheet(wb[sheet])
    meta['lore']     = LORE[i]
    levels.append(meta)

js = '// Auto-generated — do not edit manually\n'
js += f'const GAME_DATA = {json.dumps({"levels": levels}, indent=2, ensure_ascii=False)};\n'

with open('js/data.js', 'w', encoding='utf-8') as f:
    f.write(js)

print('OK  js/data.js generated')
for lv in levels:
    print(f"   Level {lv['id']} ({lv['name']}): {len(lv['problems'])} problems")
